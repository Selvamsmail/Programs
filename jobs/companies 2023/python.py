import pandas as pd
import openpyxl

# Load data from Excel files
df_2021 = pd.read_excel('data_2021.xlsx')
df_2022 = pd.read_excel('data_2022.xlsx')

# Get distinct company names from both years
distinct_companies = set(df_2021['company_name']).union(set(df_2022['company_name']))

# Create a new DataFrame for the output
output_df = pd.DataFrame({'Companies': list(distinct_companies)})

# Merge with 2021 data and mark 'Yes' or 'No' in the '2021' column
output_df = output_df.merge(df_2021[['company_name']], how='left', left_on='Companies', right_on='company_name', suffixes=('', '_2021'))
output_df['2021'] = output_df['company_name'].apply(lambda x: 'Yes' if pd.notna(x) else 'No')

# Merge with 2022 data and mark 'Yes' or 'No' in the '2022' column
output_df = output_df.merge(df_2022[['company_name']], how='left', left_on='Companies', right_on='company_name', suffixes=('', '_2022'))
output_df['2022'] = output_df['company_name_2022'].apply(lambda x: 'Yes' if pd.notna(x) else 'No')

# Drop unnecessary columns
output_df = output_df.drop(['company_name', 'company_name_2022'], axis=1)

# Set up Excel writer
excel_writer = pd.ExcelWriter('output_file.xlsx', engine='openpyxl')

# Write data to Excel
output_df.to_excel(excel_writer, index=False, sheet_name='Sheet1')

# Get the active sheet
workbook = excel_writer.book
worksheet = workbook['Sheet1']

# Define the colors
red_fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
green_fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# Apply formatting for the '2021' column
for row in worksheet.iter_rows(min_row=2, max_row=output_df.shape[0] + 1, min_col=2, max_col=2):
    for cell in row:
        if cell.value == 'Yes':
            cell.fill = green_fill
        else:
            cell.fill = red_fill

# Apply formatting for the '2022' column
for row in worksheet.iter_rows(min_row=2, max_row=output_df.shape[0] + 1, min_col=3, max_col=3):
    for cell in row:
        if cell.value == 'Yes':
            cell.fill = green_fill
        else:
            cell.fill = red_fill

# Save the Excel file
excel_writer.save()
